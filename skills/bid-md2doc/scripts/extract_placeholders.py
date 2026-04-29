#!/usr/bin/env python3
"""
占位符提取工具 - 从Word文档中提取所有占位符并生成Excel清单

功能：
1. 扫描Word文档中的所有段落文本
2. 查找所有【此处插入XX】格式的占位符
3. 生成Excel清单，供用户填写本地图片路径
4. 完全离线工作，不依赖任何API

使用方法：
    python extract_placeholders.py <word文档路径>

输出：
    占位符清单.xlsx - 包含占位符列表和待填写的图片路径列
"""

import sys
import re
from pathlib import Path
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def extract_placeholders_from_docx(docx_path: str) -> list:
    """
    从Word文档中提取所有占位符

    Args:
        docx_path: Word文档路径

    Returns:
        占位符列表，每个元素包含 {序号, 占位符原文, 材料名称, 位置信息}
    """
    try:
        doc = Document(docx_path)
    except Exception as e:
        print(f"❌ 无法打开Word文档: {e}")
        sys.exit(1)

    # 占位符正则表达式：【此处插入XX扫描件】或【此处插入XX】
    placeholder_pattern = re.compile(r'【此处插入(.+?)】')

    placeholders = []
    paragraph_num = 0

    print(f"📄 正在扫描文档: {Path(docx_path).name}")
    print("=" * 60)

    # 扫描所有段落
    for para in doc.paragraphs:
        paragraph_num += 1
        text = para.text

        # 查找占位符
        for match in placeholder_pattern.finditer(text):
            full_placeholder = match.group(0)  # 完整占位符文本
            material_name = match.group(1).strip()  # 提取材料名称

            # 清理材料名称（去除"扫描件"、"复印件"等后缀）
            clean_name = material_name
            for suffix in ['扫描件', '复印件', '正反面', '彩色', '清晰']:
                clean_name = clean_name.replace(suffix, '').strip()

            placeholder_info = {
                'index': len(placeholders) + 1,
                'placeholder': full_placeholder,
                'material_name': material_name,
                'clean_name': clean_name,
                'location': f'第 {paragraph_num} 段',
                'context': text[:50] + '...' if len(text) > 50 else text,
            }

            placeholders.append(placeholder_info)
            print(f"✓ 找到占位符 #{len(placeholders)}: {full_placeholder}")

    print("=" * 60)
    print(f"✅ 共找到 {len(placeholders)} 个占位符\n")

    return placeholders


def generate_excel_template(placeholders: list, output_path: str):
    """
    生成Excel清单模板

    Args:
        placeholders: 占位符列表
        output_path: 输出Excel文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "占位符清单"

    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30

    # 标题样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 写入标题行
    headers = ['序号', '占位符原文', '材料名称', '本地图片路径', '位置', '上下文']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 数据行样式
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    index_alignment = Alignment(horizontal="center", vertical="center")

    # 写入占位符数据
    for row_num, ph in enumerate(placeholders, 2):
        # 序号
        cell = ws.cell(row=row_num, column=1, value=ph['index'])
        cell.alignment = index_alignment

        # 占位符原文
        cell = ws.cell(row=row_num, column=2, value=ph['placeholder'])
        cell.alignment = data_alignment

        # 材料名称（清理后的）
        cell = ws.cell(row=row_num, column=3, value=ph['clean_name'])
        cell.alignment = data_alignment

        # 本地图片路径（空白，待用户填写）
        cell = ws.cell(row=row_num, column=4, value='')
        cell.alignment = data_alignment
        # 高亮提示用户需要填写
        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        # 位置
        cell = ws.cell(row=row_num, column=5, value=ph['location'])
        cell.alignment = data_alignment

        # 上下文
        cell = ws.cell(row=row_num, column=6, value=ph['context'])
        cell.alignment = data_alignment

    # 添加说明工作表
    ws_info = wb.create_sheet(title="使用说明")
    ws_info.column_dimensions['A'].width = 80

    instructions = [
        "=== 占位符替换工具使用说明 ===",
        "",
        "1. 填写说明：",
        '   - 在"本地图片路径"列中填写您电脑上对应材料的图片文件路径',
        "   - 支持格式：PNG、JPG、JPEG",
        "   - 路径示例：C:\\Users\\您的用户名\\Documents\\材料\\营业执照.png",
        "   - 如果某个占位符不需要替换，可以留空",
        "",
        "2. 运行替换脚本：",
        "   python replace_placeholders.py",
        "",
        "3. 注意事项：",
        "   - 请确保图片文件存在且可读",
        "   - 图片会被自动调整大小以适应Word文档",
        "   - 替换后会生成新文件：{原文件名}-已替换.docx",
        "   - 原始Word文档不会被修改",
        "",
        "4. 常见问题：",
        "   - 如果路径包含中文，请确保使用UTF-8编码保存Excel",
        "   - Windows路径使用反斜杠 \\ 或正斜杠 /",
        "   - 相对路径和绝对路径都支持",
        "",
        "5. 示例路径格式：",
        "   Windows: C:\\材料\\营业执照.png 或 C:/材料/营业执照.png",
        "   Linux/Mac: /home/user/材料/营业执照.png 或 ~/材料/营业执照.png",
        "   相对路径: ./材料/营业执照.png",
    ]

    for row_num, instruction in enumerate(instructions, 1):
        cell = ws_info.cell(row=row_num, column=1, value=instruction)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        if instruction.startswith("==="):
            cell.font = Font(bold=True, size=14)
        elif instruction.startswith("1.") or instruction.startswith("2.") or \
             instruction.startswith("3.") or instruction.startswith("4.") or \
             instruction.startswith("5."):
            cell.font = Font(bold=True, size=11)

    # 保存Excel文件
    try:
        wb.save(output_path)
        print(f"✅ 已生成占位符清单: {output_path}")
    except Exception as e:
        print(f"❌ 保存Excel失败: {e}")
        sys.exit(1)


def main():
    if len(sys.argv) < 2:
        print("用法: python extract_placeholders.py <Word文档路径>")
        print("示例: python extract_placeholders.py 投标文件.docx")
        sys.exit(1)

    docx_path = sys.argv[1]

    # 检查文件是否存在
    if not Path(docx_path).exists():
        print(f"❌ 文件不存在: {docx_path}")
        sys.exit(1)

    # 检查是否为Word文档
    if not docx_path.endswith('.docx'):
        print("❌ 请提供.docx格式的Word文档")
        sys.exit(1)

    print("\n" + "=" * 60)
    print("📋 占位符提取工具 v1.0")
    print("=" * 60 + "\n")

    # 提取占位符
    placeholders = extract_placeholders_from_docx(docx_path)

    if not placeholders:
        print("ℹ️  文档中没有找到占位符，无需生成清单")
        return

    # 生成Excel清单（保存在Word文档同目录）
    docx_dir = Path(docx_path).parent
    excel_path = docx_dir / "占位符清单.xlsx"

    generate_excel_template(placeholders, str(excel_path))

    print("\n" + "=" * 60)
    print("✅ 占位符提取完成！")
    print("=" * 60)
    print(f"\n📊 占位符清单: {excel_path}")
    print("\n下一步操作：")
    print("1. 打开 占位符清单.xlsx")
    print('2. 在"本地图片路径"列填写您的材料图片路径')
    print("3. 运行 python replace_placeholders.py 进行替换")
    print()


if __name__ == '__main__':
    main()
