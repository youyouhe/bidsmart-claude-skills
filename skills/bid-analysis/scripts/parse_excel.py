#!/usr/bin/env python3
"""
Excel 文件解析工具
提取 Excel 中的所有工作表数据，输出为 JSON 格式供 LLM 分析
"""
import argparse
import json
import logging
import sys
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def parse_excel(file_path: str | Path) -> dict[str, Any]:
    """
    解析 Excel 文件，提取所有工作表数据

    Args:
        file_path: Excel 文件路径

    Returns:
        包含文件信息和所有工作表数据的字典
    """
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    logger.info(f"Parsing Excel file: {file_path}")

    # 使用 data_only=True 读取公式计算后的值
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        logger.error(f"Failed to load Excel file: {e}")
        raise

    result = {
        "file_name": file_path.name,
        "file_path": str(file_path.absolute()),
        "sheet_count": len(wb.worksheets),
        "sheets": []
    }

    for sheet in wb.worksheets:
        logger.info(f"Processing sheet: {sheet.title}")

        sheet_data = {
            "name": sheet.title,
            "row_count": sheet.max_row,
            "col_count": sheet.max_column,
            "rows": []
        }

        # 提取所有行数据
        for row in sheet.iter_rows(values_only=True):
            # 将 None 转为空字符串，数字保留原值
            clean_row = []
            for cell in row:
                if cell is None:
                    clean_row.append("")
                elif isinstance(cell, (int, float)):
                    clean_row.append(cell)
                else:
                    clean_row.append(str(cell))
            sheet_data["rows"].append(clean_row)

        result["sheets"].append(sheet_data)
        logger.info(f"Sheet '{sheet.title}': {sheet.max_row} rows, {sheet.max_column} columns")

    logger.info(f"Parsed {len(result['sheets'])} sheets from {file_path.name}")
    return result


def format_as_markdown(data: dict[str, Any]) -> str:
    """
    将 Excel 数据格式化为 Markdown 表格

    Args:
        data: parse_excel 返回的数据字典

    Returns:
        Markdown 格式的文本
    """
    md_lines = [
        f"# {data['file_name']}",
        "",
        f"**文件路径**: {data['file_path']}",
        f"**工作表数量**: {data['sheet_count']}",
        ""
    ]

    for sheet in data["sheets"]:
        md_lines.append(f"## 工作表: {sheet['name']}")
        md_lines.append("")
        md_lines.append(f"**行数**: {sheet['row_count']}, **列数**: {sheet['col_count']}")
        md_lines.append("")

        if sheet["rows"]:
            # 生成 Markdown 表格
            rows = sheet["rows"]
            if len(rows) > 0:
                # 表头
                header = rows[0]
                md_lines.append("| " + " | ".join(str(cell) for cell in header) + " |")
                md_lines.append("|" + "|".join(["---"] * len(header)) + "|")

                # 数据行（最多显示前100行）
                max_display_rows = 100
                for row in rows[1:max_display_rows + 1]:
                    md_lines.append("| " + " | ".join(str(cell) for cell in row) + " |")

                if len(rows) > max_display_rows + 1:
                    md_lines.append("")
                    md_lines.append(f"*（共 {len(rows)} 行，仅显示前 {max_display_rows} 行）*")

                md_lines.append("")
        else:
            md_lines.append("*（工作表为空）*")
            md_lines.append("")

    return "\n".join(md_lines)


def main():
    parser = argparse.ArgumentParser(
        description="Parse Excel files and extract structured data"
    )
    parser.add_argument(
        "input",
        help="Input Excel file (.xlsx, .xlsm, .xls)"
    )
    parser.add_argument(
        "-o", "--output",
        help="Output JSON file path (default: <input_name>_data.json)"
    )
    parser.add_argument(
        "--markdown",
        help="Output Markdown file path (optional)"
    )
    parser.add_argument(
        "--format",
        choices=["json", "markdown", "both"],
        default="json",
        help="Output format (default: json)"
    )

    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Error: File not found: {input_path}")
        sys.exit(1)

    # 解析 Excel
    try:
        data = parse_excel(input_path)
    except Exception as e:
        print(f"Error parsing Excel: {e}")
        sys.exit(1)

    # 确定输出路径
    if args.output:
        json_output = Path(args.output)
    else:
        json_output = input_path.parent / f"{input_path.stem}_data.json"

    # 输出 JSON
    if args.format in ["json", "both"]:
        with open(json_output, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"✅ JSON output saved to: {json_output}")

    # 输出 Markdown
    if args.format in ["markdown", "both"] or args.markdown:
        if args.markdown:
            md_output = Path(args.markdown)
        else:
            md_output = input_path.parent / f"{input_path.stem}_data.md"

        md_content = format_as_markdown(data)
        with open(md_output, 'w', encoding='utf-8') as f:
            f.write(md_content)
        print(f"✅ Markdown output saved to: {md_output}")

    # 打印统计信息
    print(f"\n📊 Summary:")
    print(f"   File: {data['file_name']}")
    print(f"   Sheets: {data['sheet_count']}")
    for sheet in data["sheets"]:
        print(f"     - {sheet['name']}: {sheet['row_count']} rows × {sheet['col_count']} cols")


if __name__ == "__main__":
    main()
